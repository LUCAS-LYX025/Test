# -*- coding: utf-8 -*-

"""
@author: lucas
@Function: 定义写入文件类型
@file: writeHotToFile.py
@time: 2021/9/29 2:28 下午
"""
import os

import openpyxl
import pandas as pd

from utils.config import DATA_PATH, Config
from utils.fileOperate import fileOperate

FORMAT = ['.html', '.csv', '.json', '.xlsx']


def writeHotToFile(data, n):
    file = pd.DataFrame(data, columns=['排名', '今日热搜', '热度(单位为万)'])
    print(file)
    encoding = Config().get('general_parameters').get('charset')
    exportname = fileOperate(path=DATA_PATH).fileNaming()

    if n >= len(FORMAT):
        print("Error: 数组大小超限")
    else:
        base_path = DATA_PATH + '/' + exportname + FORMAT[n]
        if n == 0:
            return file.to_html(base_path, encoding=encoding[1])
        elif n == 1:
            return file.to_csv(base_path)
        elif n == 2:
            return file.to_json(base_path, force_ascii=False)
        elif n == 3:
            return file.to_excel(base_path)
        else:
            return "抱歉！没有找到符合的文件格式，请重新输入！"


def writeToExcel(result, sheet_name, exportname):
    os.chdir('%s' % DATA_PATH)
    inwb = openpyxl.Workbook()
    inwb.create_sheet(sheet_name, 0)
    inwb.remove(inwb.worksheets[1])
    ws = inwb.worksheets[0]
    for i in range(len(result)):
        for j in range(len(result[0])):
            ws.cell(row=i + 1, column=j + 1).value = result[i][j]
    inwb.save('%s.xlsx' % exportname)
    print("success !!!")
