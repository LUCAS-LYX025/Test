import requests
import time
import pandas as pd
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from urllib.parse import urlparse
import threading
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import socket
import ssl
import urllib3
from typing import Dict, List, Optional, Tuple
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# 禁用不安全的请求警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


class DetailedResponseTimeTester:
    def __init__(self):
        self.results = []
        self.test_count = 0
        self.running = False
        self.url_types = []  # 存储URL类型信息
        self.session = requests.Session()  # 使用Session保持连接

    def test_url_with_details(self, url: str, url_type: str, timeout: int = 10) -> Dict:
        """详细测试URL的各个阶段时间（毫秒）"""
        parsed_url = urlparse(url)
        host = parsed_url.hostname

        result = {
            'timestamp': datetime.now(),
            'url': url,
            'url_type': url_type,
            'success': False,
            'error': None,
            'status_code': None,
            'waiting_time': None,  # TTFB (Time To First Byte)
            'content_download_time': None,  # 内容下载时间
            'total_time': None,  # 总时间
            'content_length': None
        }

        overall_start = time.time()  # 记录整个测试的开始时间

        try:
            # 准备HTTP请求头
            headers = {
                'Cache-Control': 'no-cache, no-store, must-revalidate',
                'Pragma': 'no-cache',
                'Expires': '0',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                'Accept-Encoding': 'gzip, deflate, br, zstd',
                'Accept-Language': 'zh-CN,zh;q=0.9',
                'Connection': 'keep-alive',
                'Host': host,
                'Sec-Ch-Ua': '"Not;A=Brand";v="99", "Google Chrome";v="139", "Chromium";v="139"',
                'Sec-Ch-Ua-Mobile': '?0',
                'Sec-Ch-Ua-Platform': '"macOS"',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'none',
                'Sec-Fetch-User': '?1',
                'Upgrade-Insecure-Requests': '1',
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36'
            }

            cookies = {
                '__bid_n': '18e084bcdfe3168b1ff1ff',
                'cna': 'c61db3e14b004b5aba15cc48c43de357',
                '_ga': 'GA1.1.1741264619.1745808865',
                'sensorsdata2015jssdkcross': '%7B%22distinct_id%22%3A%22193ba66b54d3d7-0b251547d93b35-1e525636-1296000-193ba66b54ea9c%22%2C%22first_id%22%3A%22%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMTkzYmE2NmI1NGQzZDctMGIyNTE1NDdkOTNiMzUtMWU1MjU2MzYtMTI5NjAwMC0xOTNiYTY2YjU0ZWE5YyJ9%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%22%2C%22value%22%3A%22%22%7D%2C%22%24device_id%22%3A%22193ba66b54d3d7-0b251547d93b35-1e525636-1296000-193ba66b54ea9c%22%7D',
                '_uetvid': 'ca2c09a0678e11f0b20d3f99b003220f',
                '_fbp': 'fb.1.1754364511858.740231316197767154',
                'Hm_lvt_fd5bc3b9d2848cc07a819fc1490badd8': '1755069725',
                '_gcl_au': '1.1.1070825922.1754364507.421367491.1756979625.1756979743',
                '_ga_QTXYKYDTTR': 'GS2.1.s1757398858$o34$g1$t1757398862$j56$l0$h652625744',
                '_clck': 'afs6ir%5E2%5Efze%5E0%5E2003'
            }

            # 使用Session发送请求
            request_start = time.time()

            with self.session.get(
                    url,
                    timeout=timeout,
                    verify=False,
                    stream=True,
                    headers=headers,
                    cookies=cookies
            ) as response:
                # 1. 测量TTFB（Time To First Byte）
                # response.elapsed包含从发送请求到收到响应头的时间
                ttfb = response.elapsed.total_seconds() * 1000
                result['waiting_time'] = ttfb

                # 2. 准确测量内容下载时间
                # 下载时间应该从第一个字节开始计算，到最后一个字节结束
                first_byte_time = time.time()  # 第一个字节到达的时间

                # 读取完整内容
                content = response.content
                last_byte_time = time.time()  # 最后一个字节到达的时间

                # 计算下载时间（从第一个字节到最后一个字节）
                content_download_time = (last_byte_time - first_byte_time) * 1000
                result['content_download_time'] = content_download_time

                result['status_code'] = response.status_code
                result['success'] = True
                result['content_length'] = len(content)

            # 3. 计算总时间（从请求开始到最后一个字节）
            total_time = (time.time() - request_start) * 1000
            result['total_time'] = total_time

            # 验证时间关系：总时间 ≈ TTFB + 下载时间
            calculated_total = result['waiting_time'] + result['content_download_time']
            time_diff = abs(total_time - calculated_total)

            # 如果时间差异太大，说明测量有问题
            if time_diff > 50:  # 50ms的容忍度
                print(f"时间测量不一致: 总时间={total_time:.2f}ms, TTFB+下载={calculated_total:.2f}ms, 差异={time_diff:.2f}ms")

        except requests.exceptions.RequestException as e:
            result['error'] = str(e)
            result['total_time'] = (time.time() - overall_start) * 1000
        except Exception as e:
            result['error'] = str(e)
            result['total_time'] = (time.time() - overall_start) * 1000

        self.results.append(result)
        self.test_count += 1
        return result

    def run_continuous_test(self, urls: List[Tuple[str, str]], interval: int = 5, max_tests: Optional[int] = None):
        """连续测试多个URL的响应时间"""
        self.running = True
        count = 0

        while self.running and (max_tests is None or count < max_tests):
            for url, url_type in urls:
                if not self.running:
                    break
                self.test_url_with_details(url, url_type)
                time.sleep(1)  # 在测试不同URL之间添加短暂延迟
            count += 1
            time.sleep(interval)

    def stop_test(self):
        """停止测试"""
        self.running = False
        self.session.close()  # 关闭session

    def get_stats(self) -> Dict:
        """获取详细的统计信息（毫秒）"""
        if not self.results:
            return None

        df = pd.DataFrame(self.results)
        successful = df[df['success']]

        if successful.empty:
            return {
                'total_tests': len(df),
                'successful_tests': 0,
                'success_rate': 0
            }

        stats = {
            'total_tests': len(df),
            'successful_tests': len(successful),
            'success_rate': len(successful) / len(df) * 100,
        }

        # 为每个时间指标添加统计信息（毫秒）
        time_metrics = ['waiting_time', 'content_download_time', 'total_time']

        for metric in time_metrics:
            if metric in successful.columns:
                valid_times = successful[metric].dropna()
                if not valid_times.empty:
                    stats.update({
                        f'avg_{metric}': valid_times.mean(),
                        f'min_{metric}': valid_times.min(),
                        f'max_{metric}': valid_times.max(),
                        f'std_{metric}': valid_times.std()
                    })

        # 添加时间关系验证
        if 'waiting_time' in successful.columns and 'content_download_time' in successful.columns:
            successful['calculated_total'] = successful['waiting_time'] + successful['content_download_time']
            successful['time_diff'] = abs(successful['total_time'] - successful['calculated_total'])
            stats['avg_time_diff'] = successful['time_diff'].mean()
            stats['max_time_diff'] = successful['time_diff'].max()

        return stats

    def get_comparison_stats(self) -> Dict:
        """获取不同URL类型的对比统计信息"""
        if not self.results:
            return None

        df = pd.DataFrame(self.results)
        successful = df[df['success']]

        if successful.empty:
            return None

        # 按URL类型分组
        grouped = successful.groupby('url_type')

        stats = {}
        time_metrics = ['waiting_time', 'content_download_time', 'total_time']

        # 为每种URL类型和每个时间指标计算统计信息
        for url_type, group in grouped:
            stats[url_type] = {}
            for metric in time_metrics:
                if metric in group.columns:
                    valid_times = group[metric].dropna()
                    if not valid_times.empty:
                        stats[url_type][metric] = {
                            'avg': valid_times.mean(),
                            'min': valid_times.min(),
                            'max': valid_times.max(),
                            'std': valid_times.std()
                        }

        return stats

    def export_results(self, filename: Optional[str] = None) -> str:
        """导出结果到Excel文件，包含详细数据和统计结果"""
        if not self.results:
            return False

        if filename is None:
            domain = urlparse(self.results[0]['url']).netloc
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"detailed_response_times_{domain}_{timestamp}.xlsx"
        else:
            # 确保文件名以.xlsx结尾
            if not filename.lower().endswith('.xlsx'):
                filename += '.xlsx'

        # 创建Excel文件
        wb = Workbook()

        # 创建详细数据工作表
        ws_raw = wb.active
        ws_raw.title = "详细测试数据"

        # 添加标题行
        headers = list(self.results[0].keys())
        ws_raw.append(headers)

        # 设置标题行样式
        bold_font = Font(bold=True)
        for cell in ws_raw[1]:
            cell.font = bold_font

        # 添加数据行
        for result in self.results:
            ws_raw.append([result[key] for key in headers])

        # 创建统计结果工作表
        ws_stats = wb.create_sheet(title="统计结果")

        # 添加测试概况统计
        stats = self.get_stats()
        ws_stats.append(["测试概况统计"])
        ws_stats.append(["总测试次数", stats['total_tests']])
        ws_stats.append(["成功测试次数", stats['successful_tests']])
        ws_stats.append(["成功率", f"{stats['success_rate']:.2f}%"])
        ws_stats.append([])  # 空行

        # 添加详细时间统计标题
        ws_stats.append(["详细时间统计 (毫秒)"])
        ws_stats.append(["时间指标", "平均值(ms)", "最小值(ms)", "最大值(ms)", "标准差(ms)"])

        # 添加时间统计数据
        time_metrics = [
            ('等待响应时间(TTFB)', 'waiting_time'),
            ('内容下载时间', 'content_download_time'),
            ('总时间', 'total_time')
        ]

        for display_name, metric in time_metrics:
            if f'avg_{metric}' in stats:
                ws_stats.append([
                    display_name,
                    f"{stats[f'avg_{metric}']:.3f}",
                    f"{stats[f'min_{metric}']:.3f}",
                    f"{stats[f'max_{metric}']:.3f}",
                    f"{stats[f'std_{metric}']:.3f}"
                ])

        # 添加时间关系验证
        if 'avg_time_diff' in stats:
            ws_stats.append([])
            ws_stats.append(["时间关系验证"])
            ws_stats.append(["TTFB + 下载时间 ≈ 总时间"])
            ws_stats.append(["平均时间差异", f"{stats['avg_time_diff']:.3f} ms"])
            ws_stats.append(["最大时间差异", f"{stats['max_time_diff']:.3f} ms"])

        # 设置标题样式
        for row in [1, 6, 7, 12]:
            for cell in ws_stats[row]:
                cell.font = bold_font

        # 创建对比统计工作表（如果有多种URL类型）
        comparison_stats = self.get_comparison_stats()
        if comparison_stats and len(comparison_stats) > 1:
            ws_comparison = wb.create_sheet(title="对比统计")

            # 添加标题
            ws_comparison.append(["不同格式图片链接性能对比 (毫秒)"])

            # 添加表头
            headers = ["时间指标", "URL类型", "平均值", "最小值", "最大值", "标准差"]
            ws_comparison.append(headers)

            # 添加数据
            row_num = 3
            for metric_display, metric in time_metrics:
                for url_type in comparison_stats:
                    if metric in comparison_stats[url_type]:
                        ws_comparison.append([
                            metric_display,
                            url_type,
                            comparison_stats[url_type][metric]['avg'],
                            comparison_stats[url_type][metric]['min'],
                            comparison_stats[url_type][metric]['max'],
                            comparison_stats[url_type][metric]['std']
                        ])
                        row_num += 1
                # 添加空行分隔不同指标
                ws_comparison.append([])
                row_num += 1

            # 设置标题样式
            for cell in ws_comparison[1]:
                cell.font = bold_font
            for cell in ws_comparison[2]:
                cell.font = bold_font

            # 创建图表
            chart = LineChart()
            chart.title = "不同格式图片链接响应时间对比"
            chart.style = 13
            chart.y_axis.title = '时间 (毫秒)'
            chart.x_axis.title = '测试次数'

            # 准备图表数据
            df = pd.DataFrame(self.results)
            df = df[df['success']]  # 只使用成功的测试

            # 按URL类型分组
            url_types = df['url_type'].unique()

            # 创建时间序列图表
            for url_type in url_types:
                # 获取该URL类型的所有测试结果
                type_data = df[df['url_type'] == url_type]

                # 添加总时间数据
                data = Reference(ws_comparison, min_col=3, min_row=3, max_row=row_num - 1)
                chart.add_data(data, titles_from_data=True)

            # 设置X轴标签
            categories = Reference(ws_comparison, min_col=1, min_row=3, max_row=row_num - 1)
            chart.set_categories(categories)

            # 添加图表到工作表
            ws_comparison.add_chart(chart, "A" + str(row_num + 5))

        # 设置列宽
        for ws in wb.worksheets:
            for col in range(1, len(headers) + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 20

        # 保存文件
        wb.save(filename)
        return os.path.abspath(filename)


class DetailedResponseTimeGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("图片链接响应时间测试工具")
        self.root.geometry("1200x800")

        self.tester = DetailedResponseTimeTester()

        self.setup_ui()

    def setup_ui(self):
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # URL输入区域
        url_frame = ttk.LabelFrame(main_frame, text="测试URL（可选多种格式）", padding="5")
        url_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)

        # 格式1
        ttk.Label(url_frame, text="格式1 (原图):").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.url1_var = tk.StringVar(
            value="https://hqh5test1.zuiyouliao.com/static/img/hy_tq1_4.55e28079.png")
        url1_entry = ttk.Entry(url_frame, textvariable=self.url1_var, width=80)
        url1_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5, pady=2)

        # 格式2
        ttk.Label(url_frame, text="格式2 (中等尺寸):").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.url2_var = tk.StringVar(
            value="https://oss.plastmatch.com/zx/image/3a8d2753c19b48c2acf727237952220a.jpg?x-oss-process=image/resize,m_lfit,w_600")
        url2_entry = ttk.Entry(url_frame, textvariable=self.url2_var, width=80)
        url2_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5, pady=2)

        # 格式3
        ttk.Label(url_frame, text="格式3 (小尺寸):").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.url3_var = tk.StringVar(
            value="https://oss.plastmatch.com/zx/image/3a8d2753c19b48c2acf727237952220a.jpg?x-oss-process=image/resize,m_lfit,w_260")
        url3_entry = ttk.Entry(url_frame, textvariable=self.url3_var, width=80)
        url3_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=5, pady=2)

        # 测试参数区域
        params_frame = ttk.LabelFrame(main_frame, text="测试参数", padding="5")
        params_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)

        ttk.Label(params_frame, text="间隔时间(秒):").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.interval_var = tk.IntVar(value=5)
        interval_entry = ttk.Entry(params_frame, textvariable=self.interval_var, width=10)
        interval_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=2)

        ttk.Label(params_frame, text="最大测试次数:").grid(row=0, column=2, sticky=tk.W, pady=2)
        self.max_tests_var = tk.IntVar(value=10)
        max_tests_entry = ttk.Entry(params_frame, textvariable=self.max_tests_var, width=10)
        max_tests_entry.grid(row=0, column=3, sticky=tk.W, padx=5, pady=2)

        # 按钮区域
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.grid(row=2, column=0, columnspan=2, pady=10)

        self.start_button = ttk.Button(buttons_frame, text="开始详细测试", command=self.start_test)
        self.start_button.pack(side=tk.LEFT, padx=5)

        self.stop_button = ttk.Button(buttons_frame, text="停止测试", command=self.stop_test, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=5)

        self.export_button = ttk.Button(buttons_frame, text="导出结果", command=self.export_results)
        self.export_button.pack(side=tk.LEFT, padx=5)

        self.stats_button = ttk.Button(buttons_frame, text="显示详细统计", command=self.show_stats)
        self.stats_button.pack(side=tk.LEFT, padx=5)

        # 结果显示区域
        results_frame = ttk.LabelFrame(main_frame, text="详细测试结果", padding="5")
        results_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)

        # 创建表格形式的显示
        columns = ("时间", "URL类型", "状态", "TTFB(ms)", "下载(ms)", "总时间(ms)")
        self.tree = ttk.Treeview(results_frame, columns=columns, show="headings", height=15)

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, anchor=tk.CENTER)

        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)

        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)

    def start_test(self):
        # 获取三种格式的URL
        url1 = self.url1_var.get().strip()
        url2 = self.url2_var.get().strip()
        url3 = self.url3_var.get().strip()

        # 检查至少有一个URL被填写
        if not url1 and not url2 and not url3:
            messagebox.showerror("错误", "请至少输入一个有效的URL")
            return

        interval = self.interval_var.get()
        max_tests = self.max_tests_var.get()

        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.status_var.set("详细测试中...")

        # 清空现有结果
        for item in self.tree.get_children():
            self.tree.delete(item)

        # 准备URL列表和类型
        urls = []
        if url1:
            urls.append((url1, "原图"))
        if url2:
            urls.append((url2, "中等尺寸"))
        if url3:
            urls.append((url3, "小尺寸"))

        # 在新线程中运行测试
        self.test_thread = threading.Thread(
            target=self.tester.run_continuous_test,
            args=(urls, interval, max_tests),
            daemon=True
        )
        self.test_thread.start()

        # 启动UI更新
        self.update_results()

    def stop_test(self):
        self.tester.stop_test()
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.status_var.set("测试已停止")

    def update_results(self):
        if self.tester.running:
            # 更新树形显示
            for item in self.tree.get_children():
                self.tree.delete(item)

            # 显示最近结果（每种URL类型最多显示10条）
            displayed_results = {}
            for result in reversed(self.tester.results):
                url_type = result['url_type']
                if url_type not in displayed_results:
                    displayed_results[url_type] = []
                if len(displayed_results[url_type]) < 10:
                    displayed_results[url_type].append(result)

            # 合并所有结果并显示
            all_results = []
            for url_type in displayed_results:
                all_results.extend(displayed_results[url_type])

            # 按时间排序
            all_results.sort(key=lambda x: x['timestamp'], reverse=True)

            # 显示在表格中
            for result in all_results[:30]:  # 最多显示30条
                self.tree.insert("", tk.END, values=(
                    result['timestamp'].strftime('%H:%M:%S'),
                    result['url_type'],
                    result['status_code'] if result['success'] else 'Error',
                    f"{result['waiting_time']:.3f}" if result['waiting_time'] else 'N/A',
                    f"{result['content_download_time']:.3f}" if result['content_download_time'] else 'N/A',
                    f"{result['total_time']:.3f}" if result['total_time'] else 'N/A'
                ))

            # 每秒更新一次
            self.root.after(1000, self.update_results)

    def export_results(self):
        if not self.tester.results:
            messagebox.showwarning("警告", "没有测试结果可导出")
            return

        try:
            filename = self.tester.export_results()
            if filename:
                messagebox.showinfo("成功", f"详细结果已导出到:\n{filename}")
            else:
                messagebox.showerror("错误", "导出失败")
        except Exception as e:
            messagebox.showerror("导出错误", f"导出过程中发生错误:\n{str(e)}")

    def show_stats(self):
        if not self.tester.results:
            messagebox.showwarning("警告", "没有测试结果可显示")
            return

        stats = self.tester.get_stats()
        comparison_stats = self.tester.get_comparison_stats()

        stats_window = tk.Toplevel(self.root)
        stats_window.title("详细测试统计")
        stats_window.geometry("600x500")

        notebook = ttk.Notebook(stats_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 总体统计标签页
        overall_frame = ttk.Frame(notebook)
        notebook.add(overall_frame, text="总体统计")

        overall_text = scrolledtext.ScrolledText(overall_frame, width=70, height=20)
        overall_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        overall_text.insert(tk.END, "=== 测试概况 ===\n")
        overall_text.insert(tk.END, f"总测试次数: {stats['total_tests']}\n")
        overall_text.insert(tk.END, f"成功测试次数: {stats['successful_tests']}\n")
        overall_text.insert(tk.END, f"成功率: {stats['success_rate']:.2f}%\n\n")

        overall_text.insert(tk.END, "=== 详细时间统计 (毫秒) ===\n")

        time_metrics = [
            ('等待响应时间(TTFB)', 'waiting_time'),
            ('内容下载时间', 'content_download_time'),
            ('总时间', 'total_time')
        ]

        for display_name, metric in time_metrics:
            if f'avg_{metric}' in stats:
                overall_text.insert(tk.END, f"\n{display_name}:\n")
                overall_text.insert(tk.END, f"  平均: {stats[f'avg_{metric}']:.3f} ms\n")
                overall_text.insert(tk.END, f"  最小: {stats[f'min_{metric}']:.3f} ms\n")
                overall_text.insert(tk.END, f"  最大: {stats[f'max_{metric}']:.3f} ms\n")
                overall_text.insert(tk.END, f"  标准差: {stats[f'std_{metric}']:.3f} ms\n")

        # 添加时间关系验证
        if 'avg_time_diff' in stats:
            overall_text.insert(tk.END, f"\n=== 时间关系验证 ===\n")
            overall_text.insert(tk.END, f"TTFB + 下载时间 ≈ 总时间\n")
            overall_text.insert(tk.END, f"平均时间差异: {stats['avg_time_diff']:.3f} ms\n")
            overall_text.insert(tk.END, f"最大时间差异: {stats['max_time_diff']:.3f} ms\n")

        overall_text.config(state=tk.DISABLED)

        # 对比统计标签页（如果有多种URL类型）
        if comparison_stats and len(comparison_stats) > 1:
            comparison_frame = ttk.Frame(notebook)
            notebook.add(comparison_frame, text="对比统计")

            comparison_text = scrolledtext.ScrolledText(comparison_frame, width=70, height=20)
            comparison_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            comparison_text.insert(tk.END, "=== 不同格式图片链接性能对比 (毫秒) ===\n\n")

            for display_name, metric in time_metrics:
                comparison_text.insert(tk.END, f"{display_name}:\n")

                for url_type in comparison_stats:
                    if metric in comparison_stats[url_type]:
                        comparison_text.insert(tk.END, f"  {url_type}:\n")
                        comparison_text.insert(tk.END, f"    平均: {comparison_stats[url_type][metric]['avg']:.3f} ms\n")
                        comparison_text.insert(tk.END, f"    最小: {comparison_stats[url_type][metric]['min']:.3f} ms\n")
                        comparison_text.insert(tk.END, f"    最大: {comparison_stats[url_type][metric]['max']:.3f} ms\n")
                        comparison_text.insert(tk.END, f"    标准差: {comparison_stats[url_type][metric]['std']:.3f} ms\n")

                comparison_text.insert(tk.END, "\n")

            comparison_text.config(state=tk.DISABLED)


if __name__ == "__main__":
    root = tk.Tk()
    app = DetailedResponseTimeGUI(root)
    root.mainloop()