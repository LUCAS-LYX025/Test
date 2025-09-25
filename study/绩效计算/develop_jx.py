import pymysql
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
import logging
from typing import Dict, List, Optional


class BugStatsExporter:
    def __init__(self):
        # 禅道数据库配置
        self.db_config = {
            'host': '192.168.30.34',
            'port': 3306,
            'user': 'leiyx',
            'password': 'Leiyx#2022',
            'database': 'zentao',
            'charset': 'utf8mb4',
            'cursorclass': pymysql.cursors.DictCursor
        }

        # 配置日志
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            filename='../bug_stat_export.log'
        )

    def get_db_connection(self):
        """获取数据库连接"""
        try:
            return pymysql.connect(**self.db_config)
        except pymysql.Error as e:
            logging.error(f"数据库连接失败: {str(e)}")
            raise

    def build_query(self, product_id: int, start_date: str, end_date: str,
                    global_config: Dict) -> str:
        """构建SQL查询语句"""
        exclude_types_str = ", ".join(f"'{t}'" for t in global_config['exclude_types'])
        roles_str = ", ".join(f"'{r}'" for r in global_config['roles'])

        return f"""
        SELECT 
            CASE 
                WHEN u.role = 'qdev' THEN '前端开发'
                WHEN u.role = 'hdev' THEN '后端开发'
                WHEN u.role = 'adev' THEN 'App开发'
                ELSE '其他'
            END AS 开发角色,
            u.realname AS 开发人员,
            COALESCE(SUM(CASE WHEN combined_data.is_high_priority = 1 THEN combined_data.bug_count ELSE 0 END), 0) AS 一级超时响应次数,
            COALESCE(SUM(CASE WHEN combined_data.is_high_priority = 0 THEN combined_data.bug_count ELSE 0 END), 0) AS 普通超时响应次数,
            COALESCE(SUM(combined_data.bug_count), 0) AS 总超时响应次数,
            COALESCE(total_bugs.bug_total, 0) AS 开发总bug数,
            CASE 
                WHEN COALESCE(total_bugs.bug_total, 0) = 0 THEN '0%'
                ELSE CONCAT(ROUND(COALESCE(SUM(combined_data.bug_count), 0) / COALESCE(total_bugs.bug_total, 1) * 100, 2), '%')
            END AS 超时响应率
        FROM 
            zt_user u
        LEFT JOIN (
            -- 已解决的bug统计（按resolvedBy）
            SELECT 
                resolvedBy AS developer,
                (severity = 1 OR pri = 1) AS is_high_priority,
                COUNT(*) AS bug_count
            FROM 
                zt_bug
            WHERE 
                product = {product_id}
                AND openedDate BETWEEN '{start_date}' AND '{end_date}'
                AND resolvedDate IS NOT NULL
                AND deleted = '0'
                AND type NOT IN ({exclude_types_str})
                AND (
                    -- 高优先级：考虑周末顺延
                    ((severity = 1 OR pri = 1) AND 
                        CASE 
                            -- 如果是周五、周六、周日(6,7,1)创建的bug，顺延48小时
                            WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                            THEN TIMESTAMPDIFF(HOUR, openedDate, resolvedDate) > 72  -- 24+48小时
                            ELSE TIMESTAMPDIFF(HOUR, openedDate, resolvedDate) > 24   -- 正常24小时
                        END)
                    OR
                    -- 普通优先级：考虑周末顺延
                    ((severity != 1 AND pri != 1) AND 
                        CASE 
                            -- 如果是周五、周六、周日(6,7,1)创建的bug，顺延48小时
                            WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                            THEN TIMESTAMPDIFF(HOUR, openedDate, resolvedDate) > 120  -- 72+48小时
                            ELSE TIMESTAMPDIFF(HOUR, openedDate, resolvedDate) > 72   -- 正常72小时
                        END)
                )
            GROUP BY 
                resolvedBy, is_high_priority

            UNION ALL

            -- 未解决的bug统计（按assignedTo）
            SELECT 
                assignedTo AS developer,
                (severity = 1 OR pri = 1) AS is_high_priority,
                COUNT(*) AS bug_count
            FROM 
                zt_bug
            WHERE 
                product = {product_id}
                AND openedDate BETWEEN '{start_date}' AND '{end_date}'
                AND status = 'active'
                AND deleted = '0'
                AND type NOT IN ({exclude_types_str})
                AND (
                    -- 高优先级：考虑周末顺延
                    ((severity = 1 OR pri = 1) AND 
                        CASE 
                            -- 如果是周五、周六、周日(6,7,1)创建的bug，顺延48小时
                            WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                            THEN DATE_ADD(openedDate, INTERVAL 72 HOUR) < NOW()  -- 24+48小时
                            ELSE DATE_ADD(openedDate, INTERVAL 24 HOUR) < NOW() -- 正常24小时
                        END)
                    OR
                    -- 普通优先级：考虑周末顺延
                    ((severity != 1 AND pri != 1) AND 
                        CASE 
                            -- 如果是周五、周六、周日(6,7,1)创建的bug，顺延48小时
                            WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                            THEN DATE_ADD(openedDate, INTERVAL 120 HOUR) < NOW()  -- 72+48小时
                            ELSE DATE_ADD(openedDate, INTERVAL 72 HOUR) < NOW()   -- 正常72小时
                        END)
                )
            GROUP BY 
                assignedTo, is_high_priority
        ) AS combined_data ON u.account = combined_data.developer
        LEFT JOIN (
            -- 计算每个开发人员的总bug数（assignedTo和resolvedBy的总和）
            SELECT 
                developer,
                SUM(bug_count) AS bug_total
            FROM (
                -- 已解决的bug（按resolvedBy统计）
                SELECT 
                    resolvedBy AS developer,
                    COUNT(*) AS bug_count
                FROM 
                    zt_bug
                WHERE 
                    product = {product_id}
                    AND openedDate BETWEEN '{start_date}' AND '{end_date}'
                    AND deleted = '0'
                    AND type NOT IN ({exclude_types_str})
                GROUP BY 
                    resolvedBy

                UNION ALL

                -- 未解决的bug（按assignedTo统计）
                SELECT 
                    assignedTo AS developer,
                    COUNT(*) AS bug_count
                FROM 
                    zt_bug
                WHERE 
                    product = {product_id}
                    AND openedDate BETWEEN '{start_date}' AND '{end_date}'
                    AND status = 'active'
                    AND deleted = '0'
                    AND type NOT IN ({exclude_types_str})
                GROUP BY 
                    assignedTo
            ) AS all_bugs
            GROUP BY 
                developer
        ) AS total_bugs ON u.account = total_bugs.developer
        WHERE 
            u.role IN ({roles_str})  -- 只统计指定角色的开发人员
            AND COALESCE(total_bugs.bug_total, 0) > 0  -- 只统计有bug的开发人员
        GROUP BY 
            开发角色, u.realname, u.account, total_bugs.bug_total
        ORDER BY 
            开发角色,
            COALESCE(SUM(combined_data.bug_count), 0) DESC,
            COALESCE(SUM(CASE WHEN combined_data.is_high_priority = 1 THEN combined_data.bug_count ELSE 0 END), 0) DESC,
            COALESCE(SUM(CASE WHEN combined_data.is_high_priority = 0 THEN combined_data.bug_count ELSE 0 END), 0) DESC;
        """

    def export_to_excel(self, products_config: Dict[int, Dict], global_config: Dict) -> Optional[str]:
        """
        导出多产品数据到Excel
        :param products_config: 产品配置字典 {
            product_id: {
                'name': '产品名称',
                'start_date': '开始日期',
                'end_date': '结束日期'
            }
        }
        :param global_config: 全局配置 {
            'exclude_types': ['要排除的类型'],
            'roles': ['要统计的角色']
        }
        :return: 生成的Excel文件名
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"缺陷响应统计_{timestamp}.xlsx"

            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                for product_id, config in products_config.items():
                    # 合并项目配置和全局配置
                    query_params = {
                        'start_date': config['start_date'],
                        'end_date': config['end_date'],
                        **global_config
                    }

                    df = self.query_product_stats(product_id, query_params)

                    if df is not None and not df.empty:
                        sheet_name = self._create_sheet_name(config['name'])
                        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                        self._format_worksheet(writer.book[sheet_name], df, config)

            logging.info(f"数据已成功导出到 {excel_filename}")
            return excel_filename

        except Exception as e:
            logging.error(f"导出Excel时出错: {str(e)}")
            return None

    def query_product_stats(self, product_id: int, params: Dict) -> Optional[pd.DataFrame]:
        """查询单个产品的统计数据"""
        try:
            query = self.build_query(
                product_id=product_id,
                start_date=params['start_date'],
                end_date=params['end_date'],
                global_config=params
            )
            with self.get_db_connection() as conn:
                return pd.read_sql(query, conn)
        except Exception as e:
            logging.error(f"查询产品 {product_id} 数据时出错: {str(e)}")
            return None

    def _create_sheet_name(self, product_name: str) -> str:
        """创建合法的工作表名称"""
        invalid_chars = ':\\/?*[]'
        sheet_name = product_name[:25].strip()
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '')
        return sheet_name or "产品统计"

    def _format_worksheet(self, worksheet, df, config):
        """格式化Excel工作表"""
        # 添加产品标题（包含日期范围）
        title = f"{config['name']} ({config['start_date'][:10]}至{config['end_date'][:10]})"
        worksheet.cell(row=1, column=1, value=title)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df.shape[1])

        # 设置标题样式
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for col in range(1, df.shape[1] + 1):
            cell = worksheet.cell(row=2, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置列宽
        column_widths = {'A': 15, 'B': 20, 'C': 18, 'D': 18}
        for col_letter, width in column_widths.items():
            if ord(col_letter) - ord('A') < df.shape[1]:
                worksheet.column_dimensions[col_letter].width = width

        # 设置内容对齐
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, max_col=df.shape[1]):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 冻结表头
        worksheet.freeze_panes = 'A3'


if __name__ == "__main__":
    try:
        # 全局统一配置
        global_config = {
            'exclude_types': ['change', 'designdefect', 'increased', 'document_defect', 'design', 'interface',
                              'function', 'data', 'standard'],
            # 需求变动、设计缺陷、新增需求、文档缺陷、设计如此、界面优化、功能优化、数据问题、标准规范
            'roles': ['qdev', 'hdev', 'adev']
        }

        # 产品配置字典（每个项目单独设置日期范围）
        products_config = {
            20: {
                'name': "专塑视界APP",
                'start_date': '2025-07-30 00:00:00',
                'end_date': '2025-08-28 23:59:59'
            },
            50: {
                'name': "海外官网",
                'start_date': '2025-07-30 00:00:00',
                'end_date': '2025-08-28 23:59:59'
            },
            51: {
                'name': "海外APP项目",
                'start_date': '2025-07-30 00:00:00',
                'end_date': '2025-08-28 23:59:59'
            }
            # 可以继续添加更多项目...
        }

        exporter = BugStatsExporter()
        result_file = exporter.export_to_excel(products_config, global_config)

        if result_file:
            print(f"导出成功，文件已保存为: {result_file}")
            print("包含以下产品统计:")
            for config in products_config.values():
                print(f"- {config['name']} ({config['start_date'][:10]}至{config['end_date'][:10]})")
        else:
            print("没有数据需要导出")

    except Exception as e:
        print(f"程序执行失败: {str(e)}")
        print("详细信息请查看日志文件: bug_stat_export.log")
