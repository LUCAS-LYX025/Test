import requests
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# 原始数据
data = [
    ["低代码", "https://prezcloud-jy.zuiyouliao.com/low-code-boot/doc.html",
     "https://zcloud.zuiyouliao.com/low-code-boot/doc.html"],
    ["行情-upas", "https://zshqadminpre-jy.zuiyouliao.com/api/upas/swagger-ui.html",
     "https://zshqadmin.zuiyouliao.com/api/upas/swagger-ui.html"],
    ["行情-information", "https://zshqadminpre-jy.zuiyouliao.com/api/information/swagger-ui.html",
     "https://zshqadmin.zuiyouliao.com/api/information/swagger-ui.html"],
    ["行情-quotation", "https://zshqadminpre-jy.zuiyouliao.com/api/quotation/swagger-ui.html",
     "https://zshqadmin.zuiyouliao.com/api/quotation/swagger-ui.html"],
    ["行情-customer", "https://zshqadminpre-jy.zuiyouliao.com/api/customer/swagger-ui.html",
     "https://zshqadmin.zuiyouliao.com/api/customer/swagger-ui.html"],
    ["行情-search", "https://zshqadminpre-jy.zuiyouliao.com/api/business-search/swagger-ui.html",
     "https://zshqadmin.zuiyouliao.com/api/business-search/swagger-ui.html"],
    ["交易-land", "https://tradepre-jy.zuiyouliao.com/swagger-ui.html", "https://trade.zuiyouliao.com/swagger-ui.html"],
    ["交易-admin", "https://preadmin.zuiyouliao.com/swagger-ui.html", "https://admin.zuiyouliao.com/swagger-ui.html"],
    ["交易-crm", "无", "https://crm.zuiyouliao.com/crm/swagger-ui.html"]
]


def extract_domain(url):
    if url == "无":
        return None
    try:
        parsed = urlparse(url)
        return f"{parsed.scheme}://{parsed.netloc}"
    except:
        return None


def check_health_endpoint(domain):
    if not domain:
        return "无"
    url = f"{domain}/api/upas/actuator/health"
    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 404:
            return "通过"
        return f"失败({response.status_code})"
    except Exception as e:
        return f"错误({str(e)})"


# 创建Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "健康检查结果"

# 设置表头
headers = ["项目名称", "预发地址", "正式地址", "预发健康检查", "正式健康检查"]
ws.append(headers)

# 设置表头样式
header_font = Font(bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.alignment = header_alignment

# 填充数据
for row in data:
    project_name, pre_url, prod_url = row
    pre_domain = extract_domain(pre_url)
    prod_domain = extract_domain(prod_url)

    pre_check = check_health_endpoint(pre_domain)
    prod_check = check_health_endpoint(prod_domain)

    ws.append([project_name, pre_url, prod_url, pre_check, prod_check])

# 设置列宽
column_widths = [20, 60, 60, 15, 15]
for i, width in enumerate(column_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 设置自动换行
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='center')

# 保存Excel文件
filename = "健康检查结果.xlsx"
wb.save(filename)
print(f"结果已保存到 {filename}")